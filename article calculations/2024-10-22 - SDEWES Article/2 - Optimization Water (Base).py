# %%------------   IMPORT MODULES                         -----------------------------------------------------------> #
from main_classes.geothermal_system.base_bhe import (

    isobaricIntegral, ThermodynamicPoint,
    ReservoirProperties, baseEconomicEvaluator

)
from main_classes.constant import ARTICLE_CALCULATION_DIR
from openpyxl import load_workbook
import numpy as np
import os


# %%------------   GENERATE CALCULATION POINTS            -----------------------------------------------------------> #
n_points = 15

depth_list = np.round(np.linspace(1000, 5000, n_points), 1)
grad_list = np.round(np.linspace(30, 100, n_points), 1) / 1000
dt_HE = np.linspace(2, 20, n_points)

depth_list, grad_list, dt_HE = np.meshgrid(depth_list, grad_list, dt_HE, indexing='ij')

depth_list = np.ravel(depth_list)
grad_list = np.ravel(grad_list)
dt_HE = np.ravel(dt_HE)

table = np.stack((depth_list, grad_list, dt_HE)).T


# %%------------   IMPORT RESULTS                         -----------------------------------------------------------> #
CURRENT_DIR = os.path.join(ARTICLE_CALCULATION_DIR, "2024-10-22 - SDEWES Article")
file_path = os.path.join(CURRENT_DIR, "0 - Results", "Water (Base).xlsx")

workbook = load_workbook(filename=file_path)

sheet = workbook['Results']
n_rows = sheet.max_row - 1

headers = [cell.value for cell in sheet[1]]
data_dict = {header: np.empty(n_rows) for header in headers}
for header in headers:
    data_dict[header][:] = np.nan

n_row = 0
for row in sheet.iter_rows(min_row=2, values_only=True):

    for key, cell_value in zip(headers, row):

        data_dict[key][n_row] = cell_value

    n_row += 1

workbook.close()


# %%------------   EVALUATE OVERALL LCOH                  -----------------------------------------------------------> #
d_well = 0.15                       # [m]
time = 10 * (365 * 24 * 60 * 60)    # [years] to [s]
q_tot = 1000                        # [kW]
m_dot = 5.971                       # [kg/s]

depth_list = np.unique(data_dict['depth'])
grad_list = np.unique(data_dict['grad'])
curr_depth = depth_list[5]
curr_grad = grad_list[5]

T_rocks = 10 + curr_depth * curr_grad
input_point = ThermodynamicPoint(["Water"], [1], unit_system="MASS BASE SI")
output_point = input_point.duplicate()

res_prop = ReservoirProperties()
economic_evaluator = baseEconomicEvaluator()
integrator = isobaricIntegral([input_point, output_point])

res_prop.grad = curr_grad
integrator.solve_analytically = True

curr_indices = np.where(

    np.logical_and.reduce((

        data_dict['depth'] == curr_depth,
        data_dict['grad'] == curr_grad,
        data_dict['error'] < 0.5

    ))

)
new_shape = data_dict[headers[0]][curr_indices].shape

curr_data = {header: np.empty(new_shape) for header in headers}
for header in headers:
    curr_data[header] = data_dict[header][curr_indices]

curr_data.update({"LCOH": np.empty(new_shape), "l_horiz": np.empty(new_shape)})
curr_data["LCOH"][:] = np.nan
curr_data["l_horiz"][:] = np.nan

for i, dt_HE in enumerate(curr_data['dT_HE']):

    # <-- DEFINE INPUT POINTS ---------------------------->
    T_in = curr_data['T_down'][i] + 273.15  # [°C] to [K}
    P_in = curr_data['P_down'][i] * 1e3     # [kPa] to [Pa]
    T_rocks_K = T_rocks + 273.15            # [°C] to [K}

    input_point.set_variable("P", P_in)
    input_point.set_variable("T", T_in)
    output_point.set_variable("P", P_in)
    output_point.set_variable("T", T_rocks_K)

    # <-- EVALUATE L_HORIZ ------------------------------->
    integrator.limiting_points = (input_point, output_point)
    UA = integrator.evaluate_integral(curr_data['h_rel'][i])
    UdAs = np.pi * d_well / res_prop.evaluate_rel_resistance(times=[time], d=d_well)[0]
    l_horiz = UA / UdAs * integrator.dh_max * m_dot
    curr_data["l_horiz"][i] = l_horiz

    # <-- EVALUATE LCOH ---------------------------------->
    curr_data["LCOH"][i] = economic_evaluator.LCOx(

        useful_effects=q_tot,
        l_overall=l_horiz + curr_depth,
        d_well=d_well,
        other_costs=curr_data['c_tot'][i]

    )
