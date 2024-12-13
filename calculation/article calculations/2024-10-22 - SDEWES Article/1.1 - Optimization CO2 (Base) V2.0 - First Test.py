# %%------------   IMPORT MODULES                         -----------------------------------------------------------> #
from main_classes.geothermal_system.base_bhe import (

    isobaricIntegral, ThermodynamicPoint,
    ReservoirProperties, baseEconomicEvaluator

)
from main_classes.constant import ARTICLE_CALCULATION_DIR
from scipy.interpolate import griddata
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from tqdm import tqdm
import numpy as np
import os


# %%------------   IMPORT RESULTS                         -----------------------------------------------------------> #
CURRENT_DIR = os.path.join(ARTICLE_CALCULATION_DIR, "2024-10-22 - SDEWES Article")
file_path = os.path.join(CURRENT_DIR, "0 - Results", "CO2 - V2.0 (Base).xlsx")

workbook = load_workbook(filename=file_path)

sheet = workbook['Results']
n_rows = sheet.max_row - 1

headers = [cell.value for cell in sheet[1]]
data_dict = {header: np.empty(n_rows) for header in headers}
for header in headers:
    data_dict[header][:] = np.nan

n_row = 0
for row in sheet.iter_rows(min_row=2, values_only=True):

    for key, cell_value in zip(headers, row):

        data_dict[key][n_row] = cell_value

    n_row += 1

workbook.close()


# %%------------   EVALUATE LCOH                          -----------------------------------------------------------> #
d_well = 0.15                       # [m]
time = 10 * (365 * 24 * 60 * 60)    # [years] to [s]
q_tot = 1000                        # [kW]

depth_list = np.unique(data_dict['depth'])
grad_list = np.unique(data_dict['grad'])
h_rel_list = np.unique(data_dict['h_rel'])

curr_depth = depth_list[11]
curr_grad = grad_list[11]

curr_indices = np.where(np.logical_and.reduce((

    data_dict['depth'] == curr_depth,
    data_dict['grad'] == curr_grad,
    np.logical_not(np.isnan(data_dict['W_tot'])),

)))

curr_dict = {header: data_dict[header][curr_indices] for header in headers}
curr_h_rel_list = np.unique(curr_dict['h_rel'])


# <---- INITIALIZING TOOLS ------------------------------------------------->
input_point = ThermodynamicPoint(["Carbon Dioxide"], [1], unit_system="MASS BASE SI")
output_point = input_point.duplicate()

res_prop = ReservoirProperties()
res_prop.grad = curr_grad

economic_evaluator = baseEconomicEvaluator()
economic_evaluator.Le = 20

integrator = isobaricIntegral([input_point, output_point])
integrator.solve_analytically = False

T_rocks = 10 + curr_depth * curr_grad

# <---- REFINING RESULTS ----------------------------------------------->
n_points_fine = 99
sep_perc_list = np.logspace(-6, 0, n_points_fine+1)[:-1]
sep_perc = np.tile(sep_perc_list, (len(curr_h_rel_list), 1))

t_down_fine = np.tile(curr_dict['T_1'], (n_points_fine, 1)).T
p_down_fine = np.tile(curr_dict['P_2'], (n_points_fine, 1)).T
c_tot_fine = np.tile(curr_dict['C_tot'], (n_points_fine, 1)).T

m_dot_grid = np.tile(curr_dict['m_rete'], (n_points_fine, 1)).T
w_turb1_fine = np.tile(curr_dict['W_turb1'], (n_points_fine, 1)).T
w_turb2_fine = np.tile(curr_dict['W_turb2'], (n_points_fine, 1)).T
w_comp_fine = np.tile(curr_dict['W_comp'], (n_points_fine, 1)).T

m_dot_fine = m_dot_grid / (1 - sep_perc)
w_turb2_fine = w_turb2_fine * sep_perc / (1 - sep_perc)
w_net_fine = w_comp_fine - w_turb1_fine - w_turb2_fine


# %%------------   EVALUATE LCOH                          -----------------------------------------------------------> #
l_horiz_fine = np.empty(t_down_fine.shape)
c_well_fine = np.empty(t_down_fine.shape)
lcoh_fine = np.empty(t_down_fine.shape)

l_horiz_fine[:] = np.nan
c_well_fine[:] = np.nan
lcoh_fine[:] = np.nan

optimal_LCOH = np.inf
optimal_h_rel = np.nan
optimal_sep_perc = np.nan

pbar = tqdm(total=n_points_fine*len(curr_h_rel_list))

if not np.isnan(t_down_fine[0, 0]):

    # <-- DEFINE INPUT POINTS ---------------------------->
    T_in = t_down_fine[0, 0] + 273.15  # [°C] to [K]
    P_in = p_down_fine[0, 0] * 1e3  # [kPa] to [Pa]
    T_rocks_K = T_rocks + 273.15  # [°C] to [K]

    input_point.set_variable("P", P_in)
    input_point.set_variable("T", T_in)
    output_point.set_variable("P", P_in)
    output_point.set_variable("T", T_rocks_K)

    # <-- EVALUATE L_HORIZ ------------------------------->
    integrator.limiting_points = (input_point, output_point)
    UA = integrator.evaluate_integral(curr_h_rel_list)
    UdAs = np.pi * d_well / res_prop.evaluate_rel_resistance(times=[time], d=d_well)[0]

    for i, h_rel in enumerate(curr_h_rel_list):

        for j in range(len(sep_perc_list)):

            l_horiz = UA[i] / UdAs * integrator.dh_max * m_dot_fine[i, j]

            if l_horiz > 0:

                l_horiz_fine[i, j] = l_horiz

                # <-- EVALUATE LCOH ---------------------------------->
                lcoh_fine[i, j], c_well_fine[i, j] = economic_evaluator.LCOx(

                    useful_effects=q_tot,
                    l_overall=l_horiz + curr_depth,
                    d_well=d_well,
                    other_costs=c_tot_fine[i, j],
                    w_net_el=w_net_fine[i, j]

                )

                if lcoh_fine[i, j] < optimal_LCOH:

                    optimal_LCOH = lcoh_fine[i, j]
                    optimal_h_rel = curr_h_rel_list[i]
                    optimal_sep_perc = sep_perc[i, j]

            pbar.update(1)

pbar.close()


# %%------------   PLOT LCOH                              -----------------------------------------------------------> #
x, y = np.meshgrid(curr_h_rel_list, sep_perc_list, indexing='ij')

plt.contourf(x, y, np.log10(lcoh_fine))
plt.plot(optimal_h_rel, optimal_sep_perc, marker='*', markersize=15, color='#FFD700')
plt.colorbar()
plt.show()
