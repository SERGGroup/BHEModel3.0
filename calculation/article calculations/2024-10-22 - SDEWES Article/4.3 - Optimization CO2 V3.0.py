# %%------------   IMPORT MODULES                         -----------------------------------------------------------> #
from main_classes.geothermal_system.base_bhe import (

    isobaricIntegral, ThermodynamicPoint,
    ReservoirProperties, baseEconomicEvaluator

)
from main_classes.constant import ARTICLE_CALCULATION_DIR
from scipy.ndimage import gaussian_filter1d
from scipy.interpolate import griddata
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from tqdm import tqdm
import numpy as np
import os


# %%------------   IMPORT RESULTS                         -----------------------------------------------------------> #
CURRENT_DIR = os.path.join(ARTICLE_CALCULATION_DIR, "2024-10-22 - SDEWES Article")
file_path = os.path.join(CURRENT_DIR, "0 - Results", "1 - DH Network Calculation", "CO2 Case V3.0.xlsx")

workbook = load_workbook(filename=file_path)

sheet = workbook['Results']
n_rows = sheet.max_row - 1

headers = [cell.value for cell in sheet[1]]
data_dict = {header: np.empty(n_rows) for header in headers}
for header in headers:
    data_dict[header][:] = np.nan

n_row = 0
for row in sheet.iter_rows(min_row=2, values_only=True):

    for key, cell_value in zip(headers, row):

        data_dict[key][n_row] = cell_value

    n_row += 1

workbook.close()


# %%------------   EVALUATE LCOH FOR ALL                  -----------------------------------------------------------> #
d_well = 0.15                       # [m]
time = 10 * (365 * 24 * 60 * 60)    # [years] to [s]
q_tot = 22949                       # [kW]
n_wells = 10

# <---- INITIALIZING TOOLS ------------------------------------------------->
input_point = ThermodynamicPoint(["Carbon Dioxide"], [1], unit_system="MASS BASE SI")
output_point = input_point.duplicate()

res_prop = ReservoirProperties()
economic_evaluator = baseEconomicEvaluator()
economic_evaluator.Le = 20
economic_evaluator.ignore_well_cost = False

integrator = isobaricIntegral([input_point, output_point])
integrator.solve_analytically = False

depth_list = np.unique(data_dict['depth'])
grad_list = np.unique(data_dict['grad'])
h_rel_list = np.unique(data_dict['h_rel'])

depth_mesh, grad_mesh = np.meshgrid(depth_list, grad_list)

opt_results = np.empty(np.append(depth_mesh.shape, 6))
opt_results[:] = np.nan


pbar = tqdm(total=depth_mesh.shape[0]*depth_mesh.shape[1])

for i in range(depth_mesh.shape[0]):

    for j in range(depth_mesh.shape[1]):

        curr_depth = depth_mesh[i, j]
        curr_grad = grad_mesh[i, j]

        base_curr_indices = np.where(np.logical_and.reduce((

            data_dict['depth'] == curr_depth,
            data_dict['grad'] == curr_grad,
            np.logical_not(np.isnan(data_dict['W_tot'])),

        )))

        base_curr_dict = {header: data_dict[header][base_curr_indices] for header in headers}
        curr_T_sat_list = np.unique(base_curr_dict['T_sat'])

        optimal_LCOH = np.inf

        for T_sat in curr_T_sat_list:

            curr_indices = np.where(base_curr_dict['T_sat'] == T_sat)
            curr_dict = {header: base_curr_dict[header][curr_indices] for header in headers}
            curr_h_rel_list = np.unique(curr_dict['h_rel'])

            if len(curr_h_rel_list) > 0:

                T_rocks = 10 + curr_depth * curr_grad
                res_prop.grad = curr_grad

                # <---- REFINING RESULTS ----------------------------------------------->
                n_points_fine = 99
                sep_perc_list = np.logspace(-6, 0, n_points_fine+1)[:-1]
                sep_perc = np.tile(sep_perc_list, (len(curr_h_rel_list), 1))

                t_down_fine = np.tile(curr_dict['T_down'], (n_points_fine, 1)).T
                p_down_fine = np.tile(curr_dict['P_down'], (n_points_fine, 1)).T
                c_tot_fine = np.tile(curr_dict['C_tot'], (n_points_fine, 1)).T

                m_dot_grid = np.tile(curr_dict['m_grid'], (n_points_fine, 1)).T
                w_comp_fine = np.tile(curr_dict['W_comp'], (n_points_fine, 1)).T

                dh_turb = np.tile(curr_dict['dh_turb'], (n_points_fine, 1)).T
                dh_well = np.tile(curr_dict['dh_well'], (n_points_fine, 1)).T
                dh_grid = np.tile(curr_dict['dh_grid'], (n_points_fine, 1)).T

                m_dot_fine = m_dot_grid * dh_grid / (dh_well - dh_turb * sep_perc)
                w_turb_fine = dh_turb * sep_perc * m_dot_fine
                w_net_fine = w_comp_fine - w_turb_fine

                l_horiz_fine = np.empty(t_down_fine.shape)
                c_well_fine = np.empty(t_down_fine.shape)
                lcoh_fine = np.empty(t_down_fine.shape)

                l_horiz_fine[:] = np.nan
                c_well_fine[:] = np.nan
                lcoh_fine[:] = np.nan

                if not np.isnan(t_down_fine[0, 0]):

                    # <-- DEFINE INPUT POINTS ---------------------------->
                    T_in = t_down_fine[0, 0] + 273.15  # [°C] to [K]
                    P_in = p_down_fine[0, 0] * 1e3  # [kPa] to [Pa]
                    T_rocks_K = T_rocks + 273.15  # [°C] to [K]

                    input_point.set_variable("P", P_in)
                    input_point.set_variable("T", T_in)
                    output_point.set_variable("P", P_in)
                    output_point.set_variable("T", T_rocks_K)

                    # <-- EVALUATE L_HORIZ ------------------------------->
                    integrator.limiting_points = (input_point, output_point)
                    UA = integrator.evaluate_integral(curr_h_rel_list)
                    UdAs = np.pi * d_well / res_prop.evaluate_rel_resistance(times=[time], d=d_well)[0]

                    for n, h_rel in enumerate(curr_h_rel_list):

                        for m in range(len(sep_perc_list)):

                            l_horiz = UA[n] / UdAs * integrator.dh_max * (m_dot_fine[n, m]  / n_wells)

                            if l_horiz > 0:

                                l_horiz_fine[n, m] = l_horiz

                                # <-- EVALUATE LCOH ---------------------------------->
                                lcoh_fine[n, m], c_well_fine[n, m] = economic_evaluator.LCOx(

                                    useful_effects=q_tot  / n_wells,
                                    l_overall=l_horiz + curr_depth,
                                    d_well=d_well,
                                    other_costs=c_tot_fine[n, m]  / n_wells,
                                    w_net_el=w_net_fine[n, m] / n_wells

                                )

                                if lcoh_fine[n, m] < optimal_LCOH:

                                    optimal_LCOH = lcoh_fine[n, m]

                                    opt_results[i, j, 0] = optimal_LCOH
                                    opt_results[i, j, 1] = l_horiz
                                    opt_results[i, j, 2] = w_net_fine[n, m]
                                    opt_results[i, j, 3] = T_sat
                                    opt_results[i, j, 4] = h_rel_list[n]
                                    opt_results[i, j, 5] = sep_perc_list[m]

        pbar.update(1)

opt_lcoh = opt_results[:, :, 0]
opt_length = opt_results[:, :, 1]
opt_w_net = opt_results[:, :, 2]
opt_t_sat = opt_results[:, :, 3]
opt_h_rel = opt_results[:, :, 4]
opt_sep_perc = opt_results[:, :, 5]


pbar.close()


# %%------------   PLOT OPTIMAL LCOH                      -----------------------------------------------------------> #
n_fine = 150
method = 'cubic'
depth_min = 1000
lcoh_pos_mask = np.where(

    np.logical_and.reduce((

        np.logical_not(np.isnan(opt_lcoh)),
        opt_lcoh > 0,
        depth_mesh > depth_min,

    ))
)

lcoh_base_mask = np.where(

    np.logical_and.reduce((

        np.logical_not(np.isnan(opt_lcoh)),
        #depth_mesh * grad_mesh < 130,
        depth_mesh * grad_mesh > 75,

    ))
)

# For Negative LCOH
lcoh_neg_mask = np.where(

    np.logical_and.reduce((

        np.logical_not(np.isnan(opt_lcoh)),
        opt_lcoh < 0,
        depth_mesh > depth_min,

    ))
)
has_negative = len(lcoh_neg_mask[0])>0
# For length
length_mask = np.where(

    np.logical_and.reduce((

        np.logical_not(np.isnan(opt_length)),
        depth_mesh > depth_min,

    ))
)

depth_fine = np.linspace(np.min(depth_mesh[lcoh_pos_mask]), np.max(depth_mesh[lcoh_pos_mask]), n_fine)
grad_fine = np.linspace(np.min(grad_mesh[lcoh_pos_mask]), np.max(grad_mesh[lcoh_pos_mask]), n_fine)
depth_fine, grad_fine = np.meshgrid(depth_fine, grad_fine)

points = np.vstack((depth_mesh[lcoh_pos_mask], grad_mesh[lcoh_pos_mask])).T
LCOH_fine = griddata(points, np.log(opt_lcoh[lcoh_pos_mask]), (depth_fine, grad_fine), method=method)

if has_negative:
    points = np.vstack((depth_mesh[lcoh_neg_mask], grad_mesh[lcoh_neg_mask])).T
    LCOH_neg = griddata(points, np.log(-opt_lcoh[lcoh_neg_mask]), (depth_fine, grad_fine), method=method)

points = np.vstack((depth_mesh[length_mask], grad_mesh[length_mask])).T
l_horiz_fine = griddata(points, np.log(opt_length[length_mask]), (depth_fine, grad_fine), method=method)
LCOH_ovr_fine = griddata(points, opt_lcoh[length_mask], (depth_fine, grad_fine), method=method)
w_net_over_fine = griddata(points, opt_w_net[length_mask], (depth_fine, grad_fine), method=method)

points = np.vstack((depth_mesh[lcoh_base_mask], grad_mesh[lcoh_base_mask])).T
LCOH_fine_base = griddata(points, opt_lcoh[lcoh_base_mask], (depth_fine, grad_fine), method=method)

# Find the depth with minimum LCOH for each gradient
min_depths = []
min_gradients = []
min_LCOH = []
min_lwell = []
for i in range(n_fine):
    min_index = np.nanargmin(LCOH_fine_base[i, :])
    min_depths.append(depth_fine[i, min_index])
    min_gradients.append(grad_fine[i, min_index])
    min_LCOH.append(LCOH_fine_base[i, min_index])
    min_lwell.append(l_horiz_fine[i, min_index])

min_depths = np.array(min_depths)
min_gradients = np.array(min_gradients)
min_LCOH = np.array(min_LCOH)
min_lwell = np.array(min_lwell)

smooth_moving_average = True
window_size = 15  # Adjust window size as needed
half_window = window_size // 2

smooth_gaussian = True
sigma = 10

smoothed_depths = min_depths.copy()
smoothed_gradients = min_gradients.copy()
smoothed_min_LCOH = min_LCOH.copy()
smoothed_min_lwell = min_lwell.copy()

if smooth_moving_average:

    # Only apply moving average to the middle part
    for i in range(half_window, len(min_depths) - half_window):
        smoothed_depths[i] = np.mean(min_depths[i - half_window : i + half_window + 1])
        smoothed_gradients[i] = np.mean(min_gradients[i - half_window : i + half_window + 1])
        smoothed_min_LCOH[i] = np.mean(smoothed_min_LCOH[i - half_window : i + half_window + 1])
        smoothed_min_lwell[i] = np.mean(smoothed_min_lwell[i - half_window : i + half_window + 1])

if smooth_gaussian:

    smoothed_depths = gaussian_filter1d(smoothed_depths, sigma=sigma)
    smoothed_gradients = gaussian_filter1d(smoothed_gradients, sigma=sigma)
    smoothed_min_LCOH = gaussian_filter1d(smoothed_min_LCOH, sigma=sigma)
    smoothed_min_lwell = gaussian_filter1d(smoothed_min_lwell, sigma=sigma)

fig, axs = plt.subplots(1, 2, figsize=(12, 5))


# <-- FIRST AX ----------------------------------->
axs[0].set_title("Optimal LCOH [c€/kWh]")

contour = axs[0].contourf(grad_fine * 1e3, depth_fine/1e3, LCOH_fine, levels=25, cmap="viridis_r")

tick_values = np.array([1, 2, 5, 15, 50])
cbar = fig.colorbar(contour, ax=axs[0], orientation='vertical', fraction=0.046, pad=0.01)
cbar.set_ticks(np.log(tick_values / 100))
cbar.set_ticklabels(['{}'.format(tick) for tick in tick_values])
cbar.ax.set_title('+')

inline_values = np.array([5, 10, 15, 20, 25, 50])
contour_lines = axs[0].contour(

    grad_fine * 1e3, depth_fine/1e3, LCOH_fine, linestyles="solid",
    levels=np.log(inline_values / 100), colors='white',
    linewidths=1

)
plt.clabel(

    contour_lines, inline=True,
    fontsize=10, fmt=lambda x: f"{np.exp(x) * 100:.1f} c€/kWh",
    use_clabeltext=True

)

if has_negative:
    contour = axs[0].contourf(grad_fine * 1e3, depth_fine/1e3, LCOH_neg, levels=25, cmap="plasma")

    tick_values = np.array([1, 2, 5, 15, 50])

    cbar = fig.colorbar(contour, ax=axs[0], orientation='vertical', fraction=0.046, pad=0.1)
    cbar.set_ticks(np.log(tick_values / 100))
    cbar.set_ticklabels(['-{}'.format(tick) for tick in tick_values])
    cbar.ax.yaxis.set_ticks_position('left')
    cbar.ax.set_title('-')

    inline_values = np.array([5, 50])
    contour_lines = axs[0].contour(

        grad_fine * 1e3, depth_fine/1e3, LCOH_neg, linestyles="solid",
        levels=np.log(inline_values / 100), colors='white',
        linewidths=1

    )
    plt.clabel(

        contour_lines, inline=True,
        fontsize=10, fmt=lambda x: f"-{np.exp(x) * 100:.1f} c€/kWh",
        use_clabeltext=True

    )

# <-- SECOND AX ---------------------------------->
axs[1].set_title("Optimal Horizontal Length [km]")

contour = axs[1].contourf(grad_fine * 1e3, depth_fine/1e3, l_horiz_fine, levels=25)

tick_values = np.array([1, 2, 5])
cbar = fig.colorbar(contour, ax=axs[1])
cbar.set_ticks(np.log(tick_values*1e3))
cbar.set_ticklabels(['{}'.format(tick) for tick in tick_values])
cbar.ax.set_title('[km]')

inline_values = np.array([1, 2.5, 5])
contour_lines = axs[1].contour(

    grad_fine * 1e3, depth_fine/1e3, l_horiz_fine, linestyles="solid",
    levels=np.log(inline_values*1e3), colors='white',
    linewidths=1

)
plt.clabel(

    contour_lines, inline=True,
    fontsize=10, fmt=lambda x: f"{np.exp(x) / 1e3:.1f} km",
    use_clabeltext=True

)

for ax in axs:

    ax.set_xlabel("Geothermal Gradient [°C/km]")
    ax.set_ylabel("Depth [km]")
    ax.invert_yaxis()
    ax.plot(smoothed_gradients*1e3, smoothed_depths/1e3, linewidth=4, color='gold', alpha=1)

    lcoh_0_line = ax.contour(

        grad_fine * 1e3, depth_fine/1e3, LCOH_ovr_fine, linestyles="solid",
        levels=np.array([0]), colors='black',
        linewidths=5


    )

    plt.clabel(

        lcoh_0_line, inline=True,
        fontsize=10, fmt=lambda x: f"LCOH = 0",
        use_clabeltext=True

    )

plt.tight_layout(pad=1)

image_path = os.path.join(CURRENT_DIR, "0 - Output Plots", "CO2 (Base) V2.0 - Optimal LCOH.png")
plt.savefig(image_path, dpi=300)

plt.show()


# %%------------   PLOT OPTIMIZED DEPTH                   -----------------------------------------------------------> #
plt.plot(min_gradients*1e3, smoothed_min_LCOH*100, 'b-', label='LCOH')
plt.xlabel("Geothermal Gradient [°C/km]")
plt.ylabel("LCOH [c€/kWh]")
plt.ylim([-10, 40])
image_path = os.path.join(CURRENT_DIR, "0 - Output Plots", "CO2 (Base) V2.0 - Optimal LCOH profile.png")
plt.savefig(image_path, dpi=300)
plt.show()


# %%------------   PLOT OPTIMIZATION CONDITION            -----------------------------------------------------------> #
plt.plot(smoothed_gradients*1e3, np.exp(smoothed_min_lwell)/1000, label='l_horiz')
plt.plot(smoothed_gradients*1e3, smoothed_depths/1e3, label='depth')
plt.legend()
plt.xlabel("Geothermal Gradient [°C/km]")
plt.ylabel("length [km]")
plt.ylim([0, 6])
image_path = os.path.join(CURRENT_DIR, "0 - Output Plots", "CO2 (Base) V2.0 - Optimal depth and length.png")
plt.savefig(image_path, dpi=300)
plt.show()

