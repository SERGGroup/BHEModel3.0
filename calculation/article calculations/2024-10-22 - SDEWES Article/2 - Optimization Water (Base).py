# %%------------   IMPORT MODULES                         -----------------------------------------------------------> #
from main_classes.geothermal_system.base_bhe import (

    isobaricIntegral, ThermodynamicPoint,
    ReservoirProperties, baseEconomicEvaluator

)
from main_classes.constant import ARTICLE_CALCULATION_DIR
from scipy.ndimage import gaussian_filter1d
from scipy.interpolate import griddata
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from tqdm import tqdm
import numpy as np
import os


# %%------------   GENERATE CALCULATION POINTS            -----------------------------------------------------------> #
n_points = 15

depth_list = np.round(np.linspace(1000, 5000, n_points), 1)
grad_list = np.round(np.linspace(30, 100, n_points), 1) / 1000
dt_HE = np.logspace(-1, 1, n_points)

depth_list, grad_list, dt_HE = np.meshgrid(depth_list, grad_list, dt_HE, indexing='ij')

depth_list = np.ravel(depth_list)
grad_list = np.ravel(grad_list)
dt_HE = np.ravel(dt_HE)

table = np.stack((depth_list, grad_list, dt_HE)).T


# %%------------   IMPORT RESULTS                         -----------------------------------------------------------> #
CURRENT_DIR = os.path.join(ARTICLE_CALCULATION_DIR, "2024-10-22 - SDEWES Article")
file_path = os.path.join(CURRENT_DIR, "0 - Results", "Water (Base).xlsx")

workbook = load_workbook(filename=file_path)

sheet = workbook['Results']
n_rows = sheet.max_row - 1

headers = [cell.value for cell in sheet[1]]
data_dict = {header: np.empty(n_rows) for header in headers}
for header in headers:
    data_dict[header][:] = np.nan

n_row = 0
for row in sheet.iter_rows(min_row=2, values_only=True):

    for key, cell_value in zip(headers, row):

        data_dict[key][n_row] = cell_value

    n_row += 1

workbook.close()


# %%------------   EVALUATE LCOH                          -----------------------------------------------------------> #
d_well = 0.15                       # [m]
time = 10 * (365 * 24 * 60 * 60)    # [years] to [s]
q_tot = 1000                        # [kW]
m_dot = 5.971                       # [kg/s]

depth_list = np.unique(data_dict['depth'])
grad_list = np.unique(data_dict['grad'])

input_point = ThermodynamicPoint(["Water"], [1], unit_system="MASS BASE SI")
output_point = input_point.duplicate()

res_prop = ReservoirProperties()
economic_evaluator = baseEconomicEvaluator()
economic_evaluator.Le = 20

integrator = isobaricIntegral([input_point, output_point])
integrator.solve_analytically = False

new_shape = data_dict[headers[0]].shape
final_results = {header: np.empty(new_shape) for header in headers}
for header in headers:
    final_results[header] = data_dict[header]

final_results.update({"LCOH": np.empty(new_shape), "l_horiz": np.empty(new_shape), "c_well": np.empty(new_shape)})
final_results["LCOH"][:] = np.nan
final_results["l_horiz"][:] = np.nan
final_results["c_well"][:] = np.nan

optimal_results_shape = (len(depth_list) * len(grad_list))
optimal_results = {}

optimal_key_to_save = ["LCOH", "l_horiz", "h_rel", "c_tot", "c_well"]
for key in ['depth', 'grad'] + optimal_key_to_save:
    optimal_results.update({key: np.empty(optimal_results_shape)})
    optimal_results[key][:] = np.nan

n = 0
pbar = tqdm(total=len(depth_list) * len(grad_list), desc="Calculating LCOH")
for curr_depth in depth_list:

    for curr_grad in grad_list:

        optimal_results['depth'][n] = curr_depth
        optimal_results['grad'][n] = curr_grad

        T_rocks = 10 + curr_depth * curr_grad
        res_prop.grad = curr_grad

        curr_indices = np.where(

            np.logical_and.reduce((

                data_dict['depth'] == curr_depth,
                data_dict['grad'] == curr_grad,
                data_dict['error'] < 0.5

            ))

        )

        optimal_LCOH = np.inf

        for i, dt_HE in zip(curr_indices[0], final_results['dT_HE'][curr_indices]):

            is_zero_length = False

            # <-- DEFINE INPUT POINTS ---------------------------->
            T_in = final_results['T_down'][i] + 273.15  # [°C] to [K]
            P_in = final_results['P_down'][i] * 1e3     # [kPa] to [Pa]
            T_rocks_K = T_rocks + 273.15            # [°C] to [K]

            input_point.set_variable("P", P_in)
            input_point.set_variable("T", T_in)
            output_point.set_variable("P", P_in)
            output_point.set_variable("T", T_rocks_K)

            # <-- EVALUATE L_HORIZ ------------------------------->
            integrator.limiting_points = (input_point, output_point)
            UA = integrator.evaluate_integral(final_results['h_rel'][i])
            UdAs = np.pi * d_well / res_prop.evaluate_rel_resistance(times=[time], d=d_well)[0]
            l_horiz = UA / UdAs * integrator.dh_max * m_dot
            final_results["l_horiz"][i] = l_horiz

            # <-- EVALUATE LCOH ---------------------------------->
            final_results["LCOH"][i], final_results["c_well"][i] = economic_evaluator.LCOx(

                useful_effects=q_tot,
                l_overall=l_horiz + curr_depth,
                d_well=d_well,
                other_costs=final_results['c_tot'][i]

            )

            if final_results["LCOH"][i] < optimal_LCOH:

                optimal_LCOH = final_results["LCOH"][i]

                for key in optimal_key_to_save:

                    optimal_results[key][n] = final_results[key][i]

                optimal_LCOH = final_results["LCOH"][i]

        n += 1

        pbar.update(1)

pbar.close()


# %%------------   PLOT OPTIMAL LCOH                      -----------------------------------------------------------> #
n_fine = 250
method = 'cubic'
nan_mask = np.where(

    np.logical_and.reduce((

        np.logical_not(np.isnan(optimal_results['LCOH'])),
        optimal_results['depth'] > 1600,

    ))
)

depth_fine = np.linspace(np.min(optimal_results['depth'][nan_mask]), np.max(optimal_results['depth'][nan_mask]), n_fine)
grad_fine = np.linspace(np.min(optimal_results['grad'][nan_mask]), np.max(optimal_results['grad'][nan_mask]), n_fine)
depth_fine, grad_fine = np.meshgrid(depth_fine, grad_fine)

points = np.vstack((optimal_results['depth'][nan_mask], optimal_results['grad'][nan_mask])).T

LCOH_fine = griddata(points, np.log(optimal_results['LCOH'][nan_mask]), (depth_fine, grad_fine), method=method)
l_horiz_fine = griddata(points, np.log(optimal_results['l_horiz'][nan_mask]), (depth_fine, grad_fine), method=method)

# Find the depth with minimum LCOH for each gradient
min_depths = []
min_gradients = []
min_LCOH = []
min_lwell = []
for i in range(n_fine):
    min_index = np.nanargmin(LCOH_fine[i, :])
    min_depths.append(depth_fine[i, min_index])
    min_gradients.append(grad_fine[i, min_index])
    min_LCOH.append(LCOH_fine[i, min_index])
    min_lwell.append(l_horiz_fine[i, min_index])

min_depths = np.array(min_depths)
min_gradients = np.array(min_gradients)
min_LCOH = np.array(min_LCOH)
min_lwell = np.array(min_lwell)

smooth_moving_average = True
window_size = 15  # Adjust window size as needed
half_window = window_size // 2

smooth_gaussian = True
sigma = 10

smoothed_depths = min_depths.copy()
smoothed_gradients = min_gradients.copy()
smoothed_min_LCOH = min_LCOH.copy()
smoothed_min_lwell = min_lwell.copy()

if smooth_moving_average:

    # Only apply moving average to the middle part
    for i in range(half_window, len(min_depths) - half_window):
        smoothed_depths[i] = np.mean(min_depths[i - half_window : i + half_window + 1])
        smoothed_gradients[i] = np.mean(min_gradients[i - half_window : i + half_window + 1])
        smoothed_min_LCOH[i] = np.mean(smoothed_min_LCOH[i - half_window : i + half_window + 1])
        smoothed_min_lwell[i] = np.mean(smoothed_min_lwell[i - half_window : i + half_window + 1])

if smooth_gaussian:

    smoothed_depths = gaussian_filter1d(smoothed_depths, sigma=sigma)
    smoothed_gradients = gaussian_filter1d(smoothed_gradients, sigma=sigma)
    smoothed_min_LCOH = gaussian_filter1d(smoothed_min_LCOH, sigma=sigma)
    smoothed_min_lwell = gaussian_filter1d(smoothed_min_lwell, sigma=sigma)

fig, axs = plt.subplots(1, 2, figsize=(12, 5))


# <-- FIRST AX ----------------------------------->
axs[0].set_title("Optimal LCOH [c€/kWh]")

contour = axs[0].contourf(grad_fine * 1e3, depth_fine/1e3, LCOH_fine, levels=25, cmap="viridis")

tick_values = np.array([11, 25, 50, 100, 250])
cbar = fig.colorbar(contour, ax=axs[0])
cbar.set_ticks(np.log(tick_values / 100))
cbar.set_ticklabels(['{}'.format(tick) for tick in tick_values])
cbar.set_label('LCOH [c€/kWh]')

inline_values = np.array([15, 20, 25, 50])
contour_lines = axs[0].contour(

    grad_fine * 1e3, depth_fine/1e3, LCOH_fine, linestyles="solid",
    levels=np.log(inline_values / 100), colors='white',
    linewidths=1

)
plt.clabel(

    contour_lines, inline=True,
    fontsize=10, fmt=lambda x: f"{np.exp(x) * 100:.1f} c€/kWh",
    use_clabeltext=True

)


# <-- SECOND AX ---------------------------------->
axs[1].set_title("Optimal Horizontal Length [km]")

contour = axs[1].contourf(grad_fine * 1e3, depth_fine/1e3, l_horiz_fine, levels=25)

tick_values = np.array([0.5, 1, 2.5, 5, 10])
cbar = fig.colorbar(contour, ax=axs[1])
cbar.set_ticks(np.log(tick_values*1e3))
cbar.set_ticklabels(['{}'.format(tick) for tick in tick_values])
cbar.set_label('Horizontal Length [km]')

inline_values = np.array([0.5, 1, 2.5, 5])
contour_lines = axs[1].contour(

    grad_fine * 1e3, depth_fine/1e3, l_horiz_fine, linestyles="solid",
    levels=np.log(inline_values*1e3), colors='white',
    linewidths=1

)
plt.clabel(

    contour_lines, inline=True,
    fontsize=10, fmt=lambda x: f"{np.exp(x) / 1e3:.1f} km",
    use_clabeltext=True

)

for ax in axs:

    ax.set_xlabel("Geothermal Gradient [°C/km]")
    ax.set_ylabel("Depth [km]")
    ax.invert_yaxis()
    ax.plot(smoothed_gradients*1e3, smoothed_depths/1e3, linewidth=4, color='gold', alpha=1)

plt.tight_layout(pad=1)

# image_path = os.path.join(CURRENT_DIR, "0 - Output Plots", "Water (Base) - Optimal LCOH.png")
# plt.savefig(image_path, dpi=300)
plt.show()


# %%------------   PLOT OPTIMIZED DEPTH                   -----------------------------------------------------------> #
plt.plot(min_gradients*1e3, np.exp(min_LCOH)*100, 'b-', label='LCOH')
plt.xlabel("Geothermal Gradient [°C/km]")
plt.ylabel("LCOH [c€/kWh]")
plt.ylim([-10, 40])
image_path = os.path.join(CURRENT_DIR, "0 - Output Plots", "Water (Base) - Optimal LCOH profile.png")
plt.savefig(image_path, dpi=300)
plt.show()


# %%------------   PLOT OPTIMIZATION CONDITION            -----------------------------------------------------------> #
plt.plot(smoothed_gradients*1e3, np.exp(smoothed_min_lwell)/1000, label='l_horiz')
plt.plot(smoothed_gradients*1e3, smoothed_depths/1e3, label='depth')
plt.legend()
plt.xlabel("Geothermal Gradient [°C/km]")
plt.ylabel("length [km]")
plt.ylim([0, 6])
image_path = os.path.join(CURRENT_DIR, "0 - Output Plots", "Water (Base) - Optimal depth and length.png")
plt.savefig(image_path, dpi=300)
plt.show()

