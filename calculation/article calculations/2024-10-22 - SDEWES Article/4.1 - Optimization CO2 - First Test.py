# %%------------   IMPORT MODULES                         -----------------------------------------------------------> #
from main_classes.geothermal_system.base_bhe import (

    isobaricIntegral, ThermodynamicPoint,
    ReservoirProperties, baseEconomicEvaluator

)
from main_classes.constant import ARTICLE_CALCULATION_DIR
from scipy.interpolate import griddata
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from tqdm import tqdm
import numpy as np
import os


# %%------------   IMPORT RESULTS                         -----------------------------------------------------------> #
CURRENT_DIR = os.path.join(ARTICLE_CALCULATION_DIR, "2024-10-22 - SDEWES Article")
file_path = os.path.join(CURRENT_DIR, "0 - Results", "1 - DH Network Calculation", "CO2 Case V1.0.xlsx")

workbook = load_workbook(filename=file_path)

sheet = workbook['Foglio1']
n_rows = sheet.max_row - 1

headers = [cell.value for cell in sheet[1]]
data_dict = {header: np.empty(n_rows) for header in headers}
for header in headers:
    data_dict[header][:] = np.nan

n_row = 0
for row in sheet.iter_rows(min_row=2, values_only=True):

    for key, cell_value in zip(headers, row):

        data_dict[key][n_row] = cell_value

    n_row += 1

workbook.close()


# %%------------   EVALUATE LCOH                          -----------------------------------------------------------> #
d_well = 0.15                       # [m]
time = 10 * (365 * 24 * 60 * 60)    # [years] to [s]
q_tot = 22948                       # [kW]

depth_list = np.unique(data_dict['depth'])
grad_list = np.unique(data_dict['grad'])
h_rel_list = np.unique(data_dict['h_rel'])
t_sat_list = np.unique(data_dict['T_sat'])

curr_depth = depth_list[7]
curr_grad = grad_list[7]

curr_indices = np.where(np.logical_and.reduce((

    data_dict['depth'] == curr_depth,
    data_dict['grad'] == curr_grad,
    np.logical_not(np.isnan(data_dict['W_tot'])),

)))

curr_dict = {header: data_dict[header][curr_indices] for header in headers}

# <---- INITIALIZING TOOLS ------------------------------------------------->
input_point = ThermodynamicPoint(["Carbon Dioxide"], [1], unit_system="MASS BASE SI")
output_point = input_point.duplicate()

res_prop = ReservoirProperties()
economic_evaluator = baseEconomicEvaluator()
economic_evaluator.Le = 20
economic_evaluator.c_el = 0.12
economic_evaluator.hy = 8760 * 0.33

integrator = isobaricIntegral([input_point, output_point])
integrator.solve_analytically = False
T_rocks = 10 + curr_depth * curr_grad
res_prop.grad = curr_grad


# <---- REFINING RESULTS ----------------------------------------------->
method = "cubic"
n_points_fine = 20
h_rel_fine = 1 - np.logspace(-1.5, -0.35, n_points_fine)
t_sat_fine = np.linspace(10, 25, n_points_fine)
h_rel_fine, t_sat_fine = np.meshgrid(h_rel_fine, t_sat_fine)

points = np.vstack((curr_dict['h_rel'], curr_dict['T_sat'])).T
t_down_fine = griddata(points, curr_dict['T_down'], (h_rel_fine, t_sat_fine), method=method)
p_down_fine = griddata(points, curr_dict['P_down'], (h_rel_fine, t_sat_fine), method=method)
m_dot_fine = griddata(points, curr_dict['m_well'], (h_rel_fine, t_sat_fine), method=method)
c_tot_fine = griddata(points, curr_dict['C_tot'], (h_rel_fine, t_sat_fine), method=method)
w_net_fine = griddata(points, curr_dict['W_tot'], (h_rel_fine, t_sat_fine), method=method)

l_horiz_fine = np.empty(t_down_fine.shape)
c_well_fine = np.empty(t_down_fine.shape)
lcoh_fine = np.empty(t_down_fine.shape)

l_horiz_fine[:] = np.nan
c_well_fine[:] = np.nan
lcoh_fine[:] = np.nan

optimal_LCOH = np.inf
optimal_h_rel = np.nan
optimal_t_sat = np.nan

pbar = tqdm(total=n_points_fine**2)
for i, h_rel in enumerate(h_rel_fine):

    for j, t_sat in enumerate(t_sat_fine):

        if not np.isnan(t_down_fine[i, j]):

            # <-- DEFINE INPUT POINTS ---------------------------->
            T_in = t_down_fine[i, j] + 273.15  # [°C] to [K]
            P_in = p_down_fine[i, j] * 1e3  # [kPa] to [Pa]
            T_rocks_K = T_rocks + 273.15  # [°C] to [K]

            input_point.set_variable("P", P_in)
            input_point.set_variable("T", T_in)
            output_point.set_variable("P", P_in)
            output_point.set_variable("T", T_rocks_K)

            # <-- EVALUATE L_HORIZ ------------------------------->
            integrator.limiting_points = (input_point, output_point)
            UA = integrator.evaluate_integral(h_rel_fine[i, j])
            UdAs = np.pi * d_well / res_prop.evaluate_rel_resistance(times=[time], d=d_well)[0]
            l_horiz = UA / UdAs * integrator.dh_max * m_dot_fine[i, j]

            if l_horiz > 0:

                l_horiz_fine[i, j] = l_horiz

                # <-- EVALUATE LCOH ---------------------------------->
                lcoh_fine[i, j], c_well_fine[i, j] = economic_evaluator.LCOx(

                    useful_effects=q_tot,
                    l_overall=l_horiz + curr_depth,
                    d_well=d_well,
                    other_costs=c_tot_fine[i, j],
                    w_net_el=w_net_fine[i, j]

                )

                if lcoh_fine[i, j] < optimal_LCOH:

                    optimal_LCOH = lcoh_fine[i, j]
                    optimal_h_rel = h_rel_fine[i, j]
                    optimal_t_sat = t_sat_fine[i, j]

        pbar.update(1)

pbar.close()


# %%------------   PLOT LCOH                              -----------------------------------------------------------> #
plt.contourf(t_sat_fine, 1 - h_rel_fine, (lcoh_fine))
plt.plot(optimal_t_sat, 1-optimal_h_rel, marker='*', markersize=15, color='#FFD700')
plt.yscale("log")
plt.colorbar()
plt.show()
